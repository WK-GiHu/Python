import openpyxl as xl
from classes import Inschrijving


wb = xl.load_workbook("mosselen.xlsx")
ws = wb['ZATERDAG']


questions = [
    ('Voornaam: '),
    ('Achternaam: '),
    ('Aantal mosselen groot: '),
    ('Aantal mosselen groot voor helpers: '),
    ('Aantal mosselen klein: '),
    ('Aantal mosselen klein voor helpers: '),
    ('Aantal paardenworsten groot: '),
    ('Aantal paardenworsten groot voor helpers: '),
    ('Aantal paardenworsten klein: '),
    ('Aantal paardenworsten klein voor helpers: '),
    ('Brood: '),
    ('Betaald: ')]


def new():
    voornaam = input(questions[0])
    achternaam = input(questions[1])
    mg = input(questions[2])
    mgh = input(questions[3])
    mk = input(questions[4])
    mkh = input(questions[5])
    pg = input(questions[6])
    pgh = input(questions[7])
    pk = input(questions[8])
    pkh = input(questions[9])
    brood = input(questions[10])
    betaald = input(questions[11])
    insch1 = Inschrijving(voornaam=voornaam, achternaam=achternaam, mg=mg, mgh=mgh, mk=mk, mkh=mkh, pg=pg, pgh=pgh, pk=pk, pkh=pkh, brood=brood, betaald=betaald)
    print(f'{insch1.voornaam, insch1.achternaam}\nmosellen groot {insch1.mg}\nmosellen groot helper {insch1.mgh}\nmosellen klein {insch1.mk}\nmosellen klein helper {insch1.mkh}\npaardenworsten groot {insch1.pg}\npaardenworsten groot helper {insch1.pgh}\npaardenworsten klein {insch1.pk}\npaardenworsten klein helper {insch1.pkh}\nbrood {insch1.brood}\nbetaald {insch1.betaald}')

    if mg == "0":
        mg = None
    if mgh == "0":
        mgh = None
    if mk == "0":
        mk = None
    if mkh == "0":
        mkh = None
    if pg == "0":
        pg = None
    if pgh == "0":
        pgh = None
    if pk == "0":
        pk = None
    if pkh == "0":
        pkh = None
    if brood == "n":
        brood = None

    for cell in ws.iter_rows(min_row=5, min_col=4, max_col=5, max_row=ws.max_row - 9):
    # cell == (Column 4 - name, Column 5 - surname)
    if all((c.value is None for c in cell)):
        print('empty')

        # cell[0].row is the current Row
        row_index = cell[0].row
        column_index_to_start = 4

        for col_index, value in enumerate((voornaam, achternaam, mg, mgh, mk, mkh, pg, pgh, pk, pkh, brood, betaald), column_index_to_start):
            ws.cell(row=row_index, column=col_index).value = value
        break


def start():
    a = input('(1) Nieuwe inschrijving of (2) bijwerken: \n')
    if a == "1":
        new()
    elif a == "2":
        edit()
    else:
        print("Ongeldig")
        start()


def edit():
    pass


start()
wb.save(filename='mosselen_test1.xlsx')
wb.close()
